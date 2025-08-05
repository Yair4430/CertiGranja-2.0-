import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from V1.Plantilla import ajustar_ancho
import shutil

# Carpeta de descargas por defecto
ruta_carpeta_descargas = os.path.join(os.path.expanduser("~"), "Downloads")

def generar_resultados(datos, resultados_df, nombre_archivo_salida="resultados_certificados.xlsx", carpeta_destino=ruta_carpeta_descargas):
    # Cambia 'resultados' por 'resultados_df'
    if isinstance(resultados_df, list):
        resultados_df = pd.DataFrame(resultados_df)
    if resultados_df.empty:
        print("Error: El DataFrame 'resultados' está vacío. No se puede generar el archivo.")
        return

    # Carpeta de descargas
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    if not os.path.exists(downloads_folder):
        print("Error: No se puede encontrar la carpeta de Descargas.")
        return

    # Ruta general de salida (ya no por ficha)
    ruta_archivo_salida = os.path.join(downloads_folder, nombre_archivo_salida)

    # Agregar columna de observaciones directamente al DataFrame completo
    print("Agregando columna de observaciones...")
    # Asegurarse de que los resultados se asocien correctamente con las filas
    if len(resultados_df) == len(datos):
        datos["OBSERVACIONES"] = resultados_df["OBSERVACIONES"]
        datos["STATUS"] = resultados_df["STATUS"]
    else:
        print(f"ADVERTENCIA: El número de resultados ({len(resultados_df)}) no coincide con el número de filas de datos ({len(datos)})")
        # Asignar solo las observaciones disponibles
        for i in range(min(len(resultados_df), len(datos))):
            datos.loc[i, "OBSERVACIONES"] = resultados_df.iloc[i]["OBSERVACIONES"]

    # Guardar archivo con STATUS para colorear
    try:
        datos.to_excel(ruta_archivo_salida, index=False, engine="openpyxl")
        print(f"Archivo guardado en: {ruta_archivo_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")
        return

    # Ajustar ancho de columnas
    try:
        ajustar_ancho(ruta_archivo_salida)
    except Exception as e:
        print(f"Error al ajustar el ancho de las columnas: {e}")
        return

    # Aplicar colores y eliminar columna STATUS
    try:
        wb = load_workbook(ruta_archivo_salida)
        ws = wb.active
        
        # Aplicar colores basados en STATUS
        for idx, status in enumerate(datos["STATUS"] if "STATUS" in datos.columns else []):
            fill_color = None
            if pd.isna(status):
                continue
            if status == "EXITO":
                fill_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif status == "NOVEDAD":
                fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif status == "FALLIDO":
                fill_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif status == "ERROR DE PAGINA":
                fill_color = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            elif status == "ENLACE_ESPECIAL":  # Color blanco para tipos especiales
                fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            if fill_color:
                excel_row = idx + 2
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=excel_row, column=col).fill = fill_color
        
        # NUEVO: Encontrar y eliminar la columna STATUS
        status_column = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "STATUS":
                status_column = col
                break
        
        if status_column:
            ws.delete_cols(status_column)
            print("Columna STATUS eliminada del archivo Excel")
        
        wb.save(ruta_archivo_salida)
        print(f"Archivo coloreado, STATUS eliminado y guardado en: {ruta_archivo_salida}")
    except Exception as e:
        print(f"Error al aplicar colores o eliminar STATUS: {e}")
        return

    # Eliminar STATUS del DataFrame para uso posterior
    if "STATUS" in datos.columns:
        datos.drop(columns=["STATUS"], inplace=True)

    # Mover si hay carpeta destino
    if carpeta_destino:
        try:
            shutil.move(ruta_archivo_salida, os.path.join(carpeta_destino, nombre_archivo_salida))
            print(f"Archivo movido a: {os.path.join(carpeta_destino, nombre_archivo_salida)}")
        except Exception as e:
            print(f"Error al mover el archivo a la carpeta de destino: {e}")